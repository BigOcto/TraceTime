from __future__ import print_function
from json_file_editor import update_json_file
from json_file_editor import task_configure_info
from json_file_editor import TEST_JSON_FILE_PATH

import os
import subprocess
import time

import openpyxl
import xlrd
import xlsxwriter

import threading

# Collect launch times array
collect_data = []

# User command options
trace_configure_info = {'trace_times': "",
                        'trace_version_name': "",
                        'device_name': "",
                        }

apks = [
    {'name': "qiyi", 'pkg': "com.qiyi.video", 'welcome': ".WelcomeActivity", 'main': "MainActivity"},
    # {'name': "qqlive", 'pkg': "com.tencent.qqlive", 'welcome': ".ona.activity.WelcomeActivity", 'main': "HomeActivity"},
    # {'name': "letv", 'pkg': "com.letv.android.client", 'welcome': ".activity.SplashActivity", 'main': "MainActivity"},
    # {'name': "youku", 'pkg': "com.youku.phone", 'welcome': ".ActivityWelcome", 'main': "HomePageActivity"}
]


class CommandDispatcher(object):
    def __init__(self):
        self._command = None
        self._args = None

    def get_command(self, options, args):
        self._args = args
        options_dict = vars(options)
        trace_times = options_dict.get('trace_times')
        trace_version_pre = options_dict.get('trace_version_pre')
        device_name = options_dict.get('device_name')
        trace_mode = options_dict.get('trace_mode')
        trace_configure_info['trace_times'] = trace_times

        if trace_mode == 'launch_time':
            if adb_app_version(apks[0].get('pkg')) == '0.0':
                exit()
            if trace_version_pre == '0.0':
                trace_configure_info['trace_version_name'] = adb_app_version(apks[0].get('pkg'))
            else:
                trace_configure_info['trace_version_name'] = trace_version_pre + "_" + adb_app_version(
                    apks[0].get('pkg'))
            run_command_am_start(apks[0].get('pkg') + "/" + apks[0].get('welcome'), trace_times, "launch_time.xlsx",
                                 trace_configure_info.get('trace_version_name'))

        elif trace_mode == 'launch_task_time':
            if adb_app_version(apks[0].get('pkg')) == '0.0':
                exit()
            if trace_version_pre == '0.0':
                trace_configure_info['trace_version_name'] = adb_app_version(apks[0].get('pkg'))
            else:
                trace_configure_info['trace_version_name'] = trace_version_pre + "_" + adb_app_version(
                    apks[0].get('pkg'))
            loop_time = trace_times
            for key in task_configure_info:
                while trace_times > 0:
                    times = loop_time / 10
                    update_json_file("")
                    run_command_adb_push_file(TEST_JSON_FILE_PATH, '/sdcard/TaskTest.json')
                    run_command_am_start(apks[0].get('pkg') + "/" + apks[0].get('welcome'), times,
                                         "launch_task_time.xlsx",
                                         trace_configure_info.get('trace_version_name') + "_" + task_configure_info.get(
                                             key))
                    update_json_file(task_configure_info.get(key))
                    run_command_adb_push_file(TEST_JSON_FILE_PATH, '/sdcard/TaskTest.json')
                    run_command_am_start(apks[0].get('pkg') + "/" + apks[0].get('welcome'), times,
                                         'launch_task_time.xlsx', task_configure_info.get(key))
                    trace_times -= times


        elif trace_mode == 'thread_time':
            if adb_app_version(apks[0].get('pkg')) == '0.0':
                exit()
            if trace_version_pre == '0.0':
                trace_configure_info['trace_version_name'] = adb_app_version(apks[0].get('pkg'))
            else:
                trace_configure_info['trace_version_name'] = trace_version_pre + "_" + adb_app_version(
                    apks[0].get('pkg'))
            run_command_am_start_and_save_log(apks[0].get('pkg') + "/" + apks[0].get('welcome'), "adb_logcat.txt",
                                              apks[0].get('pkg'))

        else:
            device_name_list = run_command_get_devices()
            print(device_name_list)
            run_command_am_qiyi_and_qqlive(device_name_list, trace_times)


def run_command_get_devices():
    cmd = 'adb devices'
    result = subprocess.check_output(cmd.split())
    device_list = handle_get_devices_output(result)
    return device_list


def run_command_am_qiyi_and_qqlive(device_list, count):
    if count <= 0:
        return

    if len(device_list) - 1 == 1:
        stop_app('com.qiyi.video')
        cmd_qiyi = 'adb shell am start com.qiyi.video/.WelcomeActivity'
        subprocess.check_output(cmd_qiyi.split())
        count -= 1
        time.sleep(5)
        stop_app('com.qiyi.video')
        time.sleep(5)

    else:
        if len(device_list) - 1 >= 2:
            count -= 1
            stop_app_with_name('com.qiyi.video', device_list[0])
            stop_app_with_name('com.tencent.qqlive', device_list[1])
            # stop_app_with_name('com.qiyi.video', device_list[1])

            t_qiyi = threading.Thread(target=thread_run_command_start_am,
                                      args=(device_list[0], 'com.qiyi.video/.WelcomeActivity'))
            t_qqlive = threading.Thread(target=thread_run_command_start_am,
                                        args=(device_list[1], 'com.tencent.qqlive/.ona.activity.WelcomeActivity'))
            # t_qqlive = threading.Thread(target=thread_run_command_start_am,
            #                             args=(device_list[1], 'com.qiyi.video/.WelcomeActivity'))
            t_qiyi.start()
            t_qqlive.start()
            time.sleep(10)

        else:
            print("no device error!!!!!!")
            return
    run_command_am_qiyi_and_qqlive(device_list, count)


def thread_run_command_start_am(device_name, activity_name):
    cmd_qiyi = 'adb -s ' + device_name + ' shell am start ' + activity_name
    subprocess.check_output(cmd_qiyi.split())


def run_command_am_start(activity, count, excel_file_name, column_name):
    if count <= 0:
        write_excel(collect_data, apks[0].get('name') + "_" + excel_file_name, column_name)
        collect_data[:] = []
        return
    stop_app(apks[0].get('pkg'))
    print("Activity start times " + str(count))
    cmd = 'adb shell am start -W ' + activity
    result = subprocess.check_output(cmd.split())
    handle_trace_time_output(result)
    count -= 1
    time.sleep(5)
    run_command_am_start(activity, count, excel_file_name, column_name)


def run_command_am_start_and_save_log(activity_name, file_name, package):
    cmd_qiyi = 'adb shell am start ' + activity_name
    subprocess.check_output(cmd_qiyi.split())
    # log_thread = adbNewThread()
    # log_thread.start()
    time.sleep(5)
    stop_app(package)
    # log_thread.stop
    with open(file_name) as ins:
        array_thread_name = []
        array_time = []
        for line in ins:
            if 'JobManagerUtils_time' in line:
                thread_time = line.replace(" ", "")
                thread_time_re = thread_time.split(":")
                length = len(thread_time_re)
                name_re = thread_time_re[length - 2].split(",")
                print("Thread name : " + name_re[0])
                array_thread_name.append(name_re[0])
                print("Cost time : " + thread_time_re[length - 1])
                array_time.append(thread_time_re[length - 1])
        write_excel(array_thread_name, "trace_thread_time.xlsx", "thread")
        write_excel(array_time, "trace_thread_time.xlsx", "time")


def run_command_adb_push_file(source_file_path, target_file_path):
    cmd_push = 'adb push ' + source_file_path + ' ' + target_file_path
    subprocess.check_output(cmd_push.split())
    print('Push ' + source_file_path + ' to ' + target_file_path)


def adb_app_version(pkg):
    cmd = 'adb shell dumpsys package ' + pkg + ' | grep versionName'
    result = subprocess.check_output(cmd.split())
    result = result.replace(' ', '')
    result_list = result.split()
    if len(result_list) == 0:
        print(pkg + ' apk not found !!')
        return '0.0'
    else:
        print("apk version :" + str(result_list[-1]))
        return str(result_list[-1])


def stop_app(package):
    cmd = 'adb shell am force-stop ' + package
    result = subprocess.check_output(cmd.split())
    print('stop app :' + str(result.split('\r\n')))


def stop_app_with_name(package, device_name):
    cmd = 'adb -s ' + device_name + 'shell am force-stop ' + package
    subprocess.check_output(cmd.split())


def handle_trace_time_output(result):
    result = result.replace(' ', '')
    result_list = result.split()
    for re in result_list:
        print(re)
        time_re = re.split(":")
        if 'TotalTime' in time_re:
            print(time_re[-1])
            collect_data.append(time_re[-1])


def handle_get_devices_output(result):
    pre = "List of devices attached"
    pre_result = pre.replace(' ', '')
    pre_length = len(pre_result)
    result_replace = result.replace(' ', '')
    result_length = len(result_replace)
    device_and_name = result_replace[pre_length: result_length]
    device_and_name_list = device_and_name.split('device')
    return device_and_name_list


def write_excel(spend_time_list, file_name, version):
    if os.path.exists(file_name):
        update_excel(file_name, version, spend_time_list)
    else:
        create_excel(file_name)
        update_excel(file_name, version, spend_time_list)


def create_excel(excel_file):
    workbook = xlsxwriter.Workbook(excel_file)
    # row = 3
    # for t in spend_time:
    #     worksheet.write(row, 0, t)
    #     row += 1
    workbook.close()


def update_excel(excel_file, version, spend_time_list):
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.get_sheet_by_name('Sheet1')

    col_exist_number, row_exist_length = read_excel(excel_file, version)
    is_exist_col = False
    if col_exist_number != 0 and row_exist_length != 0:
        is_exist_col = True
        for i, value in enumerate(spend_time_list):
            worksheet.cell(row=i + row_exist_length + 1, column=col_exist_number + 1).value = value
    else:
        cul_count = read_first_row_size(excel_file) + 1
        worksheet.cell(row=1, column=cul_count).value = version
        for i, statN in enumerate(spend_time_list):
            if version == "time":
                worksheet.cell(row=i + 2, column=cul_count).value = int(statN)
            else:
                worksheet.cell(row=i + 2, column=cul_count).value = statN
    if is_exist_col:
        print('Write exit col ' + version + " size: " + str(collect_data) + ' success!')
    else:
        print('Write not exit col ' + version + " size: " + str(collect_data) + ' success!')
    workbook.save(excel_file)


def read_excel(excel_file, version):
    workbook = xlrd.open_workbook(excel_file)
    worksheet = workbook.sheet_by_name('Sheet1')
    first_row_list = []
    try:
        first_row_list = worksheet.row_values(0)
        print('First row list size : ' + str(len(first_row_list)))
    except Exception as e:
        print(str(e))
    col_exist_number = -1
    for col in first_row_list:
        col_exist_number += 1
        if version == col:
            col_exist_list = worksheet.col_values(0)
            list_length = len(col_exist_list)
            while list_length > 0:
                if col_exist_list[list_length - 1] is None or col_exist_list[list_length - 1] == '':
                    list_length -= 1
                else:
                    print(version + " col exist, col : " + str(col_exist_number) + ", row size : " + str(list_length))
                    break
            return col_exist_number, list_length

    print(version + " col not exist, write normal")
    return col_exist_number, 0


def read_first_row_size(excel_file):
    workbook = xlrd.open_workbook(excel_file)
    worksheet = workbook.sheet_by_name('Sheet1')
    first_row_list = []
    try:
        first_row_list = worksheet.row_values(0)
    except Exception as e:
        print(str(e))
    count = 0
    for col in first_row_list:
        count += 1

    print("cul size : " + str(count))
    return count


class adbNewThread(threading.Thread):
    def __init__(self):
        self.stopped = False
        threading.Thread.__init__(self)

    def run(self):
        cmd_save_log = 'adb logcat > ' + "log.txt"
        subprocess.check_output(cmd_save_log.split())

    def stop(self):
        self.stopped = True
